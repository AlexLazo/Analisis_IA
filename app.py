# -*- coding: utf-8 -*-
"""
Dashboard Admin API - Backend Flask
====================================
API REST completa para gestión de análisis 5W con evaluación IA.

Endpoints:
- GET  /api/analisis              - Listar análisis
- POST /api/analisis              - Crear análisis
- GET  /api/analisis/:id          - Obtener análisis específico
- PUT  /api/analisis/:id          - Actualizar análisis
- DELETE /api/analisis/:id        - Eliminar análisis
- POST /api/analisis/:id/evaluar  - Evaluar con IA
- GET  /api/estadisticas          - Estadísticas globales
- GET  /api/metricas              - Métricas temporales
- POST /api/importar              - Importar desde Excel
"""
from flask import Flask, jsonify, request, render_template, send_from_directory, send_file
from flask_cors import CORS
from pathlib import Path
import pandas as pd
from datetime import datetime
import traceback
import threading
import time
import os

import database
from database import (
    init_database,
    AnalisisDB,
    EvaluacionDB,
    MetricasDB
)
from ai_evaluator import AIEvaluator
from ai_evaluator_groq import AIEvaluatorGroq


app = Flask(__name__,
            static_folder='static',
            template_folder='templates')
CORS(app)  # Permitir CORS para desarrollo

# Inicializar base de datos
init_database()

# Seleccionar evaluador según variable de entorno AI_BACKEND
# AI_BACKEND=groq  → usa Groq API (recomendado para Railway/cloud)
# AI_BACKEND=ollama (o no definida) → usa Ollama local
_backend = os.environ.get("AI_BACKEND", "ollama").lower()

if _backend == "groq":
    print("🤖 Backend IA: Groq API")
    evaluator = AIEvaluatorGroq(model=os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant"))
else:
    print("🤖 Backend IA: Ollama local")
    evaluator = AIEvaluator(model="llama3.1:8b", timeout=180)


# ============================================================================
# RUTAS WEB (Frontend)
# ============================================================================

@app.route('/')
def index():
    """Dashboard principal"""
    return render_template('dashboard.html')


# ============================================================================
# API ENDPOINTS - Análisis 5W
# ============================================================================

@app.route('/api/analisis', methods=['GET'])
def listar_analisis():
    """Lista análisis con paginación y filtros (categoría, tipo_5w, mes)"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')  # Formato: YYYY-MM
        nivel = request.args.get('nivel')  # Nivel de análisis: Básico/Intermedio/Sostenible
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        # Construir query con filtros dinámicos
        with database.get_db() as conn:
            cursor = conn.cursor()
            query = "SELECT * FROM analisis_5w WHERE 1=1"
            params = []
            
            if categoria:
                query += " AND categoria = ?"
                params.append(categoria)
            
            if tipo_5w:
                query += " AND tipo_5w = ?"
                params.append(tipo_5w)
            
            if mes:
                query += " AND strftime('%Y-%m', fecha) = ?"
                params.append(mes)
            
            if nivel:
                # Hacer join con evaluaciones para filtrar por nivel
                query = query.replace("FROM analisis_5w WHERE", "FROM analisis_5w a JOIN evaluaciones e ON a.id = e.analisis_id WHERE")
                query += " AND e.nivel_analisis = ?"
                params.append(nivel)
            
            query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
            params.extend([limit, offset])
            
            cursor.execute(query, params)
            analisis = [dict(row) for row in cursor.fetchall()]
            
            # Contar total con mismos filtros
            count_query = "SELECT COUNT(*) as total FROM analisis_5w WHERE 1=1"
            count_params = []
            if categoria:
                count_query += " AND categoria = ?"
                count_params.append(categoria)
            if tipo_5w:
                count_query += " AND tipo_5w = ?"
                count_params.append(tipo_5w)
            if mes:
                count_query += " AND strftime('%Y-%m', fecha) = ?"
                count_params.append(mes)
            if nivel:
                count_query = count_query.replace("FROM analisis_5w WHERE", "FROM analisis_5w a JOIN evaluaciones e ON a.id = e.analisis_id WHERE")
                count_query += " AND e.nivel_analisis = ?"
                count_params.append(nivel)
            
            cursor.execute(count_query, count_params)
            total = cursor.fetchone()['total']
        
        # Agregar evaluaciones a cada análisis
        for a in analisis:
            evaluacion = EvaluacionDB.obtener_evaluacion(a['id'])
            a['evaluacion'] = evaluacion
        
        return jsonify({
            'success': True,
            'data': analisis,
            'total': total,
            'limit': limit,
            'offset': offset
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/analisis', methods=['POST'])
def crear_analisis():
    """Crea un nuevo análisis 5W"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        required = ['categoria', 'ruta', 'problema', 'por_que_1', 
                   'por_que_2', 'por_que_3', 'causa_raiz', 'plan_accion']
        
        for field in required:
            if field not in data or not data[field]:
                return jsonify({
                    'success': False,
                    'error': f'Campo requerido: {field}'
                }), 400
        
        # Crear análisis
        analisis_id = AnalisisDB.crear_analisis(data)
        
        # Auto-evaluar si se solicita
        if data.get('auto_evaluar', False):
            analisis = AnalisisDB.obtener_analisis(analisis_id)
            evaluacion = evaluator.evaluar(analisis, analisis['categoria'])
            
            eval_dict = {
                'score': evaluacion.score,
                'grade': evaluacion.grade,
                'chain_coherence': evaluacion.chain_coherence,
                'root_cause_alignment': evaluacion.root_cause_alignment,
                'action_plan_alignment': evaluacion.action_plan_alignment,
                'strengths': evaluacion.strengths,
                'weaknesses': evaluacion.weaknesses,
                'suggestions': evaluacion.suggestions,
                'tips': evaluacion.tips,
                'detail': evaluacion.detail
            }
            
            EvaluacionDB.crear_evaluacion(analisis_id, eval_dict)
            
            # Actualizar métricas
            MetricasDB.actualizar_metricas_diarias()
        
        return jsonify({
            'success': True,
            'data': {'id': analisis_id},
            'message': 'Análisis creado exitosamente'
        }), 201
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/analisis/<int:analisis_id>', methods=['GET'])
def obtener_analisis(analisis_id):
    """Obtiene un análisis específico con su evaluación"""
    try:
        analisis = AnalisisDB.obtener_analisis(analisis_id)
        
        if not analisis:
            return jsonify({
                'success': False,
                'error': 'Análisis no encontrado'
            }), 404
        
        evaluacion = EvaluacionDB.obtener_evaluacion(analisis_id)
        analisis['evaluacion'] = evaluacion
        
        return jsonify({
            'success': True,
            'data': analisis
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/analisis/<int:analisis_id>', methods=['PUT'])
def actualizar_analisis(analisis_id):
    """Actualiza un análisis existente"""
    try:
        data = request.get_json()
        
        success = AnalisisDB.actualizar_analisis(analisis_id, data)
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Análisis no encontrado'
            }), 404
        
        return jsonify({
            'success': True,
            'message': 'Análisis actualizado exitosamente'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/analisis/<int:analisis_id>', methods=['DELETE'])
def eliminar_analisis(analisis_id):
    """Elimina un análisis"""
    try:
        success = AnalisisDB.eliminar_analisis(analisis_id)
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Análisis no encontrado'
            }), 404
        
        return jsonify({
            'success': True,
            'message': 'Análisis eliminado exitosamente'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ENDPOINTS - Evaluaciones
# ============================================================================

@app.route('/api/analisis/<int:analisis_id>/evaluar', methods=['POST'])
def evaluar_analisis(analisis_id):
    """Evalúa un análisis con IA"""
    try:
        analisis = AnalisisDB.obtener_analisis(analisis_id)
        
        if not analisis:
            return jsonify({
                'success': False,
                'error': 'Análisis no encontrado'
            }), 404
        
        # Evaluar con IA
        evaluacion = evaluator.evaluar(analisis, analisis['categoria'])
        
        # Guardar evaluación
        eval_dict = {
            'score': evaluacion.score,
            'grade': evaluacion.grade,
            'nivel_analisis': evaluacion.nivel_analisis,
            'chain_coherence': evaluacion.chain_coherence,
            'root_cause_alignment': evaluacion.root_cause_alignment,
            'action_plan_alignment': evaluacion.action_plan_alignment,
            'strengths': evaluacion.strengths,
            'weaknesses': evaluacion.weaknesses,
            'suggestions': evaluacion.suggestions,
            'tips': evaluacion.tips,
            'detail': evaluacion.detail
        }
        
        eval_id = EvaluacionDB.crear_evaluacion(analisis_id, eval_dict)
        
        # Actualizar métricas
        MetricasDB.actualizar_metricas_diarias()
        
        return jsonify({
            'success': True,
            'data': {
                'id': eval_id,
                'evaluacion': eval_dict
            },
            'message': 'Evaluación completada'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/evaluaciones/recientes', methods=['GET'])
def evaluaciones_recientes():
    """Lista evaluaciones más recientes"""
    try:
        limit = int(request.args.get('limit', 20))
        evaluaciones = EvaluacionDB.listar_evaluaciones_recientes(limit)
        
        return jsonify({
            'success': True,
            'data': evaluaciones
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ENDPOINTS - Estadísticas y Métricas
# ============================================================================

@app.route('/api/estadisticas', methods=['GET'])
def obtener_estadisticas():
    """Obtiene estadísticas globales o filtradas por categoría, tipo_5w, mes"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')  # Formato: YYYY-MM
        nivel = request.args.get('nivel')
        
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            # Construir WHERE clause dinámico
            where_clauses = ["1=1"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if mes:
                where_clauses.append("strftime('%Y-%m', a.fecha) = ?")
                params.append(mes)
            
            if nivel:
                where_clauses.append("e.nivel_analisis = ?")
                params.append(nivel)
            
            where_sql = " AND ".join(where_clauses)
            
            # Estadísticas básicas
            cursor.execute(f"""
                SELECT 
                    COUNT(DISTINCT a.id) as total_analisis,
                    COUNT(DISTINCT CASE WHEN e.id IS NOT NULL THEN a.id END) as evaluados,
                    AVG(CASE WHEN e.score IS NOT NULL THEN e.score END) as promedio_score,
                    AVG(CASE WHEN e.chain_coherence IS NOT NULL THEN e.chain_coherence END) as coherencia_media,
                    AVG(CASE WHEN e.root_cause_alignment IS NOT NULL THEN e.root_cause_alignment END) as causa_media,
                    AVG(CASE WHEN e.action_plan_alignment IS NOT NULL THEN e.action_plan_alignment END) as plan_media
                FROM analisis_5w a
                LEFT JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
            """, params)
            
            row = cursor.fetchone()
            
            # Distribución por categoría (con filtros aplicados)
            cursor.execute(f"""
                SELECT 
                    a.categoria,
                    COUNT(*) as count,
                    AVG(CASE WHEN e.score IS NOT NULL THEN e.score END) as avg_score
                FROM analisis_5w a
                LEFT JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY a.categoria
                ORDER BY count DESC
            """, params)
            
            categorias_stats = [
                {
                    'categoria': r['categoria'],
                    'count': r['count'],
                    'avg_score': round(r['avg_score'], 1) if r['avg_score'] else 0
                }
                for r in cursor.fetchall()
            ]
            
            # Distribución por grades (con filtros)
            cursor.execute(f"""
                SELECT 
                    e.grade,
                    COUNT(*) as count
                FROM analisis_5w a
                INNER JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY e.grade
                ORDER BY 
                    CASE e.grade
                        WHEN 'A+' THEN 1
                        WHEN 'A' THEN 2
                        WHEN 'B' THEN 3
                        WHEN 'C' THEN 4
                        WHEN 'D' THEN 5
                        WHEN 'F' THEN 6
                        ELSE 7
                    END
            """, params)
            
            grades_stats = [{'grade': r['grade'], 'count': r['count']} for r in cursor.fetchall()]
            
            stats = {
                'total': row['total_analisis'] or 0,
                'evaluados': row['evaluados'] or 0,
                'pendientes': (row['total_analisis'] or 0) - (row['evaluados'] or 0),
                'promedio_score': round(row['promedio_score'], 1) if row['promedio_score'] else 0,
                'coherencia_media': round(row['coherencia_media'], 1) if row['coherencia_media'] else 0,
                'causa_media': round(row['causa_media'], 1) if row['causa_media'] else 0,
                'plan_media': round(row['plan_media'], 1) if row['plan_media'] else 0,
                'por_categoria': categorias_stats,
                'por_grade': grades_stats,
                'filtros_activos': {
                    'categoria': categoria,
                    'tipo_5w': tipo_5w,
                    'mes': mes,
                    'nivel': nivel
                }
            }
        
        return jsonify({
            'success': True,
            'data': stats
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/categorias', methods=['GET'])
def obtener_categorias():
    """Obtiene lista de categorías únicas con conteo"""
    try:
        with database.get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT categoria, COUNT(*) as count 
                FROM analisis_5w 
                GROUP BY categoria 
                ORDER BY count DESC
            """)
            categorias = [{'nombre': row['categoria'], 'count': row['count']} for row in cursor.fetchall()]
        
        return jsonify({
            'success': True,
            'data': categorias
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/tipos-5w', methods=['GET'])
def obtener_tipos_5w():
    """Obtiene lista de tipos de 5W únicos (Indicador de Distribución Impactado)"""
    try:
        with database.get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT tipo_5w, COUNT(*) as count 
                FROM analisis_5w 
                WHERE tipo_5w IS NOT NULL AND tipo_5w != ''
                GROUP BY tipo_5w 
                ORDER BY count DESC
            """)
            tipos = [{'valor': row['tipo_5w'], 'nombre': row['tipo_5w'], 'count': row['count']} for row in cursor.fetchall()]
        
        return jsonify({
            'success': True,
            'data': tipos
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/meses', methods=['GET'])
def obtener_meses():
    """Obtiene lista de meses únicos con conteo de análisis"""
    try:
        with database.get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    strftime('%Y-%m', fecha) as mes,
                    COUNT(*) as count
                FROM analisis_5w 
                WHERE fecha IS NOT NULL AND fecha != ''
                GROUP BY mes
                ORDER BY mes DESC
            """)
            meses = []
            for row in cursor.fetchall():
                if row['mes']:
                    # Formatear el mes para mostrar (ej: "2025-01" -> "Enero 2025")
                    year, month = row['mes'].split('-')
                    nombres_meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
                    mes_nombre = f"{nombres_meses[int(month)-1]} {year}"
                    meses.append({
                        'valor': row['mes'],  # Valor para filtro (2025-01)
                        'nombre': mes_nombre,  # Nombre para mostrar (Enero 2025)
                        'count': row['count']
                    })
        
        return jsonify({
            'success': True,
            'data': meses
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/niveles', methods=['GET'])
def obtener_niveles():
    """Obtiene lista de niveles de análisis únicos con conteo"""
    try:
        with database.get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    e.nivel_analisis,
                    COUNT(*) as count
                FROM evaluaciones e
                WHERE e.nivel_analisis IS NOT NULL AND e.nivel_analisis != ''
                GROUP BY e.nivel_analisis
                ORDER BY 
                    CASE e.nivel_analisis
                        WHEN 'Excelente' THEN 1
                        WHEN 'Bueno' THEN 2
                        WHEN 'Intermedio' THEN 3
                        WHEN 'Básico' THEN 4
                        WHEN 'Crítico' THEN 5
                        ELSE 6
                    END
            """)
            niveles = []
            for row in cursor.fetchall():
                niveles.append({
                    'valor': row['nivel_analisis'],
                    'nombre': row['nivel_analisis'],
                    'count': row['count']
                })
        
        return jsonify({
            'success': True,
            'data': niveles
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/metricas', methods=['GET'])
def obtener_metricas():
    """Obtiene métricas temporales para gráficos"""
    try:
        dias = int(request.args.get('dias', 30))
        metricas = MetricasDB.obtener_metricas_periodo(dias)
        
        return jsonify({
            'success': True,
            'data': metricas
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/metricas/desglose', methods=['GET'])
def obtener_desglose_metricas():
    """Obtiene desglose detallado de métricas por categoría con filtros opcionales"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')  # Formato: YYYY-MM
        nivel = request.args.get('nivel')
        
        import database
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            # Construir WHERE clause dinámico
            where_clauses = ["e.chain_coherence IS NOT NULL"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if mes:
                where_clauses.append("strftime('%Y-%m', a.fecha) = ?")
                params.append(mes)
            
            if nivel:
                where_clauses.append("e.nivel_analisis = ?")
                params.append(nivel)
            
            where_sql = " AND ".join(where_clauses)
            
            # Obtener promedios de coherencia, causa raíz y plan de acción por categoría
            cursor.execute(f"""
                SELECT 
                    a.categoria,
                    AVG(e.chain_coherence) as avg_coherencia,
                    AVG(e.root_cause_alignment) as avg_causa_raiz,
                    AVG(e.action_plan_alignment) as avg_plan_accion,
                    COUNT(*) as total
                FROM analisis_5w a
                JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY a.categoria
                ORDER BY a.categoria
            """, params)
            
            results = cursor.fetchall()
            desglose = []
            
            for row in results:
                desglose.append({
                    'categoria': row['categoria'],
                    'coherencia': round(row['avg_coherencia'] or 0, 2),
                    'causa_raiz': round(row['avg_causa_raiz'] or 0, 2),
                    'plan_accion': round(row['avg_plan_accion'] or 0, 2),
                    'total': row['total']
                })
            
        return jsonify({
            'success': True,
            'data': desglose
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ENDPOINTS - Gráficas Adicionales
# ============================================================================

@app.route('/api/graficas/tendencia-temporal', methods=['GET'])
def obtener_tendencia_temporal():
    """Obtiene análisis agrupados por mes para gráfica de tendencia temporal"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        nivel = request.args.get('nivel')
        
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            where_clauses = ["a.fecha IS NOT NULL"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if nivel:
                where_clauses.append("e.nivel_analisis = ?")
                params.append(nivel)
            
            where_sql = " AND ".join(where_clauses)
            
            # Obtener datos por mes
            cursor.execute(f"""
                SELECT 
                    strftime('%Y-%m', a.fecha) as mes,
                    COUNT(*) as total,
                    COUNT(e.id) as evaluados,
                    AVG(CASE WHEN e.score IS NOT NULL THEN e.score END) as score_promedio
                FROM analisis_5w a
                LEFT JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY strftime('%Y-%m', a.fecha)
                ORDER BY mes DESC
                LIMIT 12
            """, params)
            
            results = cursor.fetchall()
            tendencia = []
            
            for row in results:
                # Convertir mes a nombre legible
                year, month = row['mes'].split('-')
                meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                               'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
                mes_nombre = f"{meses_nombres[int(month)-1]} {year}"
                
                tendencia.append({
                    'mes': row['mes'],
                    'mes_nombre': mes_nombre,
                    'total': row['total'],
                    'evaluados': row['evaluados'],
                    'score_promedio': round(row['score_promedio'] or 0, 1)
                })
            
            # Revertir para que quede cronológico
            tendencia.reverse()
        
        return jsonify({
            'success': True,
            'data': tendencia
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/graficas/distribucion-niveles', methods=['GET'])
def obtener_distribucion_niveles():
    """Obtiene distribución de análisis por nivel (Básico/Intermedio/Sostenible)"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')
        
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            where_clauses = ["e.nivel_analisis IS NOT NULL"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if mes:
                where_clauses.append("strftime('%Y-%m', a.fecha) = ?")
                params.append(mes)
            
            where_sql = " AND ".join(where_clauses)
            
            cursor.execute(f"""
                SELECT 
                    e.nivel_analisis as nivel,
                    COUNT(*) as count,
                    AVG(e.score) as score_promedio
                FROM analisis_5w a
                INNER JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY e.nivel_analisis
                ORDER BY 
                    CASE e.nivel_analisis
                        WHEN 'Excelente' THEN 1
                        WHEN 'Bueno' THEN 2
                        WHEN 'Intermedio' THEN 3
                        WHEN 'Básico' THEN 4
                        WHEN 'Crítico' THEN 5
                        ELSE 6
                    END
            """, params)
            
            results = cursor.fetchall()
            niveles = []
            
            for row in results:
                niveles.append({
                    'nivel': row['nivel'],
                    'count': row['count'],
                    'score_promedio': round(row['score_promedio'] or 0, 1)
                })
        
        return jsonify({
            'success': True,
            'data': niveles
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/graficas/top-rutas', methods=['GET'])
def obtener_top_rutas():
    """Obtiene las rutas con más problemas O los mejores/peores análisis individuales"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')
        orden = request.args.get('orden', 'mas')  # 'mas' o 'menos'
        limit = int(request.args.get('limit', 10))
        tipo = request.args.get('tipo', 'agregado')  # 'agregado' o 'individual'
        
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            where_clauses = ["1=1"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if mes:
                where_clauses.append("strftime('%Y-%m', a.fecha) = ?")
                params.append(mes)
            
            where_sql = " AND ".join(where_clauses)
            order_by = "DESC" if orden == 'mas' else "ASC"
            
            # Si es tipo 'individual', devolver análisis individuales con score
            if tipo == 'individual':
                cursor.execute(f"""
                    SELECT 
                        a.id,
                        a.ruta,
                        a.categoria,
                        a.problema,
                        e.score,
                        e.grade,
                        e.nivel_analisis
                    FROM analisis_5w a
                    INNER JOIN evaluaciones e ON a.id = e.analisis_id
                    WHERE {where_sql} AND e.score IS NOT NULL
                    ORDER BY e.score {order_by}
                    LIMIT ?
                """, params + [limit])
                
                results = cursor.fetchall()
                analisis = []
                
                for row in results:
                    analisis.append({
                        'id': row['id'],
                        'ruta': row['ruta'],
                        'categoria': row['categoria'],
                        'problema': row['problema'][:100] + '...' if len(row['problema']) > 100 else row['problema'],
                        'score': round(row['score'], 1),
                        'grade': row['grade'],
                        'nivel': row['nivel_analisis']
                    })
                
                return jsonify({
                    'success': True,
                    'data': analisis
                })
            
            # Modo agregado (comportamiento original)
            cursor.execute(f"""
                SELECT 
                    a.ruta,
                    COUNT(*) as problemas,
                    AVG(CASE WHEN e.score IS NOT NULL THEN e.score END) as score_promedio
                FROM analisis_5w a
                LEFT JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
                GROUP BY a.ruta
                HAVING COUNT(*) > 0
                ORDER BY problemas {order_by}
                LIMIT ?
            """, params + [limit])
            
            results = cursor.fetchall()
            rutas = []
            
            for row in results:
                rutas.append({
                    'ruta': row['ruta'],
                    'problemas': row['problemas'],
                    'score_promedio': round(row['score_promedio'] or 0, 1)
                })
        
        return jsonify({
            'success': True,
            'data': rutas
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/graficas/metricas-promedio', methods=['GET'])
def obtener_metricas_promedio():
    """Obtiene el promedio general de las 3 métricas principales"""
    try:
        categoria = request.args.get('categoria')
        tipo_5w = request.args.get('tipo_5w')
        mes = request.args.get('mes')
        nivel = request.args.get('nivel')
        
        with database.get_db() as conn:
            cursor = conn.cursor()
            
            where_clauses = ["e.score IS NOT NULL"]
            params = []
            
            if categoria:
                where_clauses.append("a.categoria = ?")
                params.append(categoria)
            
            if tipo_5w:
                where_clauses.append("a.tipo_5w = ?")
                params.append(tipo_5w)
            
            if mes:
                where_clauses.append("strftime('%Y-%m', a.fecha) = ?")
                params.append(mes)
            
            if nivel:
                where_clauses.append("e.nivel_analisis = ?")
                params.append(nivel)
            
            where_sql = " AND ".join(where_clauses)
            
            cursor.execute(f"""
                SELECT 
                    AVG(e.chain_coherence) as coherencia,
                    AVG(e.root_cause_alignment) as causa_raiz,
                    AVG(e.action_plan_alignment) as plan_accion,
                    COUNT(*) as total_evaluaciones
                FROM analisis_5w a
                INNER JOIN evaluaciones e ON a.id = e.analisis_id
                WHERE {where_sql}
            """, params)
            
            row = cursor.fetchone()
            
            metricas = {
                'coherencia': round(row['coherencia'] or 0, 1),
                'causa_raiz': round(row['causa_raiz'] or 0, 1),
                'plan_accion': round(row['plan_accion'] or 0, 1),
                'total_evaluaciones': row['total_evaluaciones']
            }
        
        return jsonify({
            'success': True,
            'data': metricas
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# API ENDPOINTS - Importación
# ============================================================================

def evaluar_analisis_background(analisis_ids):
    """Evalúa análisis en segundo plano sin bloquear la respuesta HTTP"""
    print(f"🤖 Iniciando evaluación en background de {len(analisis_ids)} análisis...")
    evaluados = 0
    errores = 0
    timeouts = 0
    
    for idx, analisis_id in enumerate(analisis_ids, 1):
        try:
            print(f"\n📋 [{idx}/{len(analisis_ids)}] Evaluando análisis ID: {analisis_id}")
            
            analisis = AnalisisDB.obtener_analisis(analisis_id)
            if analisis:
                # Evaluar (con reintentos automáticos)
                evaluacion = evaluator.evaluar(analisis, analisis.get('categoria', 'GENERAL'))
                
                # Si la evaluación fue con fallback (error), contar como timeout
                if "Timeout" in evaluacion.detail or "Error" in evaluacion.detail:
                    timeouts += 1
                    print(f"⚠️ Evaluación con fallback para análisis {analisis_id}")
                
                eval_dict = {
                    'score': evaluacion.score,
                    'grade': evaluacion.grade,
                    'chain_coherence': evaluacion.chain_coherence,
                    'root_cause_alignment': evaluacion.root_cause_alignment,
                    'action_plan_alignment': evaluacion.action_plan_alignment,
                    'strengths': evaluacion.strengths,
                    'weaknesses': evaluacion.weaknesses,
                    'suggestions': evaluacion.suggestions,
                    'tips': evaluacion.tips,
                    'detail': evaluacion.detail
                }
                
                EvaluacionDB.crear_evaluacion(analisis_id, eval_dict)
                evaluados += 1
                
                if evaluados % 5 == 0:
                    print(f"✅ Progreso: {evaluados}/{len(analisis_ids)} evaluados ({errores} errores, {timeouts} timeouts)")
                    # Actualizar métricas parcialmente
                    MetricasDB.actualizar_metricas_diarias()
            
            # Pausa adaptativa (más larga si hubo timeouts recientes)
            pausa = 0.5 if timeouts > 3 else 0.2
            time.sleep(pausa)
            
        except Exception as e:
            print(f"❌ Error crítico evaluando análisis {analisis_id}: {str(e)[:200]}")
            errores += 1
            # Pausa más larga después de error crítico
            time.sleep(1)
    
    # Actualizar métricas al final
    MetricasDB.actualizar_metricas_diarias()
    print(f"\n🎉 Evaluación background completada:")
    print(f"   ✅ Exitosos: {evaluados}")
    print(f"   ⚠️ Timeouts/Fallback: {timeouts}")
    print(f"   ❌ Errores críticos: {errores}")

@app.route('/api/importar', methods=['POST'])
def importar_excel():
    """Importa análisis desde Excel subido por el usuario"""
    try:
        # Verificar si hay archivo en la petición
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se encontró archivo en la petición'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'El archivo debe ser Excel (.xlsx o .xls)'
            }), 400
        
        # Leer parámetros del formulario
        categoria = request.form.get('categoria')
        limite = request.form.get('limite')
        auto_evaluar = request.form.get('auto_evaluar') == 'on'  # El checkbox envía 'on' cuando está marcado
        
        print(f"⚙️ Parámetros de importación:")
        print(f"  Categoría: {categoria or 'Todas'}")
        print(f"  Límite: {limite or 'Sin límite'}")
        print(f"  Auto-evaluar: {'Sí' if auto_evaluar else 'No'}")
        
        # Leer Excel desde memoria
        df = pd.read_excel(file)
        
        print(f"📊 Excel cargado: {len(df)} filas")
        print(f"📋 Columnas encontradas: {list(df.columns)}")
        
        # Filtrar por categoría si se especifica
        if categoria:
            col_cat = None
            for col in df.columns:
                if 'categoria' in col.lower():
                    col_cat = col
                    break
            if col_cat:
                df = df[df[col_cat] == categoria]
                print(f"🔍 Filtrado por categoría '{categoria}': {len(df)} filas")
        
        # Limitar registros si se especifica
        if limite:
            df = df.head(int(limite))
            print(f"✂️ Limitado a {limite} registros")
        
        importados = 0
        errores = []
        
        # Mapeo flexible de columnas — soporta formato antiguo (con ":") y nuevo (sin ":")
        def _col(*candidatos):
            for c in candidatos:
                if c in df.columns:
                    return c
            return None

        col_area   = _col('Área:', 'Área', 'Area:', 'Area')
        col_tipo_5w = _col('Indicador de Distribución Impactado:', 'KPI Impactado', 'KPI', 'Indicador')
        col_ruta   = _col('Código de Ruta Impactada:1', 'Código de Ruta Impactada:', 'Equipo Impactado', 'Ruta')
        col_fecha  = _col('Fecha:', 'Fecha')
        col_pq1    = _col('1° ¿Por qué?', '1ero ¿Por qué?')
        col_pq2    = _col('2° ¿Por qué?', '2do ¿Por qué?')
        col_pq3    = _col('3° ¿Por qué?1', '3° ¿Por qué?', '3ro ¿Por qué?')
        col_pq4    = _col('4° ¿Por qué?1', '4° ¿Por qué?', '4to ¿Por qué?')
        col_pq5    = _col('5° ¿Por qué?1', '5° ¿Por qué?', '5to ¿Por qué?')
        col_causa  = _col('Causa Raíz - ¿Qué ocasionó la desviación?', 'Causa Raíz', 'Causa raiz')
        col_plan   = _col('Plan de Acción - ¿Cómo corregir la causa raíz?1', 'Plan de Acción - ¿Cómo corregir la causa raíz?', 'Plan de Acción', 'Plan de Accion')
        
        print(f"📋 Mapeo de columnas:")
        print(f"  Área: {col_area}")
        print(f"  Tipo 5W (Indicador): {col_tipo_5w}")
        print(f"  Ruta: {col_ruta}")
        print(f"  Fecha: {col_fecha}")
        print(f"  1° Por qué: {col_pq1}")
        print(f"  2° Por qué: {col_pq2}")
        print(f"  3° Por qué: {col_pq3}")
        print(f"  4° Por qué: {col_pq4}")
        print(f"  5° Por qué: {col_pq5}")
        print(f"  Causa Raíz: {col_causa}")
        print(f"  Plan Acción: {col_plan}")
        
        importados = 0
        errores = []
        skipped = 0
        duplicados = 0  # 🔥 Contador de duplicados detectados
        ids_importados = []  # 🔥 Lista para recolectar IDs de análisis importados
        
        for idx, row in df.iterrows():
            try:
                # Extraer valores
                pq1_val = str(row[col_pq1]) if col_pq1 and col_pq1 in row and pd.notna(row[col_pq1]) else ''
                pq2_val = str(row[col_pq2]) if col_pq2 and col_pq2 in row and pd.notna(row[col_pq2]) else ''
                pq3_val = str(row[col_pq3]) if col_pq3 and col_pq3 in row and pd.notna(row[col_pq3]) else ''
                pq4_val = str(row[col_pq4]) if col_pq4 and col_pq4 in row and pd.notna(row[col_pq4]) else ''
                pq5_val = str(row[col_pq5]) if col_pq5 and col_pq5 in row and pd.notna(row[col_pq5]) else ''
                causa_val = str(row[col_causa]) if col_causa and col_causa in row and pd.notna(row[col_causa]) else ''
                plan_val = str(row[col_plan]) if col_plan and col_plan in row and pd.notna(row[col_plan]) else ''
                
                # Usar el 1° Por qué como problema inicial
                problema_text = pq1_val if pq1_val else 'Análisis 5 Porqués'
                
                analisis_data = {
                    'categoria': str(row[col_area]) if col_area and col_area in row and pd.notna(row[col_area]) else 'GENERAL',
                    'tipo_5w': str(row[col_tipo_5w]) if col_tipo_5w and col_tipo_5w in row and pd.notna(row[col_tipo_5w]) else None,
                    'ruta': str(row[col_ruta]) if col_ruta and col_ruta in row and pd.notna(row[col_ruta]) else f'R{idx:04d}',
                    'fecha': str(row[col_fecha]) if col_fecha and col_fecha in row and pd.notna(row[col_fecha]) else '',
                    'problema': problema_text,
                    'por_que_1': pq1_val,
                    'por_que_2': pq2_val,
                    'por_que_3': pq3_val,
                    'por_que_4': pq4_val if pq4_val else pq5_val,
                    'causa_raiz': causa_val,
                    'plan_accion': plan_val
                }
                
                # Validar campos mínimos (solo verificar que no estén vacíos)
                if not pq1_val.strip() or pq1_val.strip().lower() in ['nan', 'none', '']:
                    skipped += 1
                    if skipped <= 3:
                        print(f"⏭️ Fila {idx}: Sin 1° Por qué válido")
                    continue
                    
                if not causa_val.strip() or causa_val.strip().lower() in ['nan', 'none', '']:
                    skipped += 1
                    if skipped <= 3:
                        print(f"⏭️ Fila {idx}: Sin Causa Raíz válida")
                    continue
                    
                if not plan_val.strip() or plan_val.strip().lower() in ['nan', 'none', '']:
                    skipped += 1
                    if skipped <= 3:
                        print(f"⏭️ Fila {idx}: Sin Plan de Acción válido")
                    continue
                
                # Crear análisis (ahora con detección de duplicados)
                id_antes = AnalisisDB.crear_analisis(analisis_data)
                
                # Si el ID ya existía, es un duplicado
                if id_antes in ids_importados:
                    duplicados += 1
                else:
                    ids_importados.append(id_antes)
                    importados += 1
                    
                if importados % 100 == 0:
                    print(f"✅ Importados {importados} registros...")
            
            except Exception as e:
                error_msg = f"Fila {idx}: {str(e)}"
                errores.append(error_msg)
                print(f"❌ {error_msg}")
        
        print(f"🎉 Importación completada: {importados} de {len(df)} registros")
        print(f"⏭️ Registros saltados: {skipped}")
        print(f"🔄 Duplicados detectados: {duplicados}")
        print(f"❌ Errores: {len(errores)}")
        
        # Actualizar métricas
        MetricasDB.actualizar_metricas_diarias()
        
        # 🔥 Lanzar evaluación en background si está habilitado
        mensaje_respuesta = f'✅ {importados} análisis importados.'
        if duplicados > 0:
            mensaje_respuesta += f' ⚠️ {duplicados} duplicados omitidos.'
        
        sin_evaluar = importados
        evaluando_background = False
        
        if auto_evaluar and len(ids_importados) > 0:
            print(f"🚀 Lanzando evaluación en background de {len(ids_importados)} análisis...")
            thread = threading.Thread(target=evaluar_analisis_background, args=(ids_importados,))
            thread.daemon = True  # El thread no bloqueará el cierre del servidor
            thread.start()
            mensaje_respuesta = f'✅ {importados} análisis importados. ⚡ Evaluación con IA en progreso en segundo plano... Refresca la página en unos minutos para ver los resultados.'
            sin_evaluar = 0
            evaluando_background = True
        else:
            mensaje_respuesta = f'✅ {importados} análisis importados. Usa el botón "🤖 Evaluar Todos Pendientes" para evaluarlos con IA.'
        
        return jsonify({
            'success': True,
            'data': {
                'importados': importados,
                'duplicados': duplicados,  # 🔥 Nuevo campo
                'total_procesados': len(df),
                'errores': errores[:10],  # Solo primeros 10 errores para no saturar
                'total_errores': len(errores),
                'sin_evaluar': sin_evaluar,
                'evaluando_background': evaluando_background
            },
            'message': mensaje_respuesta
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# Utilidades
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'success': True,
        'status': 'running',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/exportar-excel', methods=['GET'])
def exportar_excel():
    """Exportar análisis a Excel con todos los datos"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Obtener filtros
        categoria = request.args.get('categoria', '')
        tipo_5w = request.args.get('tipo_5w', '')
        mes = request.args.get('mes', '')
        nivel = request.args.get('nivel', '')
        
        # Obtener datos con consulta SQL directa
        import sqlite3
        import json
        
        conn = sqlite3.connect(str(database.DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Construir query con filtros
        query = """
            SELECT 
                a.*,
                e.score, e.grade, e.chain_coherence, e.root_cause_alignment, 
                e.action_plan_alignment, e.strengths, e.weaknesses, e.suggestions,
                e.tips, e.detail, e.nivel_analisis
            FROM analisis_5w a
            LEFT JOIN evaluaciones e ON a.id = e.analisis_id
            WHERE 1=1
        """
        params = []
        
        if categoria:
            query += " AND a.categoria = ?"
            params.append(categoria)
        if tipo_5w:
            query += " AND a.tipo_5w = ?"
            params.append(tipo_5w)
        if mes:
            query += " AND strftime('%Y-%m', a.fecha) = ?"
            params.append(mes)
        if nivel:
            query += " AND e.nivel_analisis = ?"
            params.append(nivel)
        
        query += " ORDER BY a.id DESC LIMIT 10000"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        analisis_list = []
        for row in rows:
            analisis = dict(row)
            # Parsear JSON de feedback IA
            if analisis.get('strengths'):
                try:
                    analisis['strengths'] = json.loads(analisis['strengths'])
                except:
                    analisis['strengths'] = []
            else:
                analisis['strengths'] = []
                
            if analisis.get('weaknesses'):
                try:
                    analisis['weaknesses'] = json.loads(analisis['weaknesses'])
                except:
                    analisis['weaknesses'] = []
            else:
                analisis['weaknesses'] = []
                
            if analisis.get('suggestions'):
                try:
                    analisis['suggestions'] = json.loads(analisis['suggestions'])
                except:
                    analisis['suggestions'] = []
            else:
                analisis['suggestions'] = []
                
            if analisis.get('tips'):
                try:
                    analisis['tips'] = json.loads(analisis['tips'])
                except:
                    analisis['tips'] = []
            else:
                analisis['tips'] = []
                
            analisis_list.append(analisis)
        
        conn.close()
        
        if not analisis_list:
            return jsonify({'success': False, 'error': 'No hay datos para exportar'}), 404
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis 5W"
        
        # Definir headers
        headers = [
            'ID', 'Categoría', 'Tipo 5W', 'Nivel', 'Ruta', 'Fecha',
            'Score', 'Grade', 'Coherencia', 'Causa Raíz %', 'Plan Acción %',
            'Problema', 'Por qué 1', 'Por qué 2', 'Por qué 3', 'Por qué 4', 'Por qué 5',
            'Causa Raíz Completa', 'Plan Completo',
            'Fortalezas', 'Debilidades', 'Sugerencias', 'Tips IA', 'Detalle IA'
        ]
        
        # Estilo de headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos
        for row_idx, analisis in enumerate(analisis_list, 2):
            ws.cell(row=row_idx, column=1, value=analisis.get('id'))
            ws.cell(row=row_idx, column=2, value=analisis.get('categoria'))
            ws.cell(row=row_idx, column=3, value=analisis.get('tipo_5w') or 'N/A')
            ws.cell(row=row_idx, column=4, value=analisis.get('nivel_analisis') or 'Sin evaluar')
            ws.cell(row=row_idx, column=5, value=analisis.get('ruta'))
            ws.cell(row=row_idx, column=6, value=analisis.get('fecha') or 'N/A')
            ws.cell(row=row_idx, column=7, value=round(analisis.get('score', 0) or 0, 1))
            ws.cell(row=row_idx, column=8, value=analisis.get('grade') or 'N/A')
            ws.cell(row=row_idx, column=9, value=round(analisis.get('chain_coherence', 0) or 0, 0))
            ws.cell(row=row_idx, column=10, value=round(analisis.get('root_cause_alignment', 0) or 0, 0))
            ws.cell(row=row_idx, column=11, value=round(analisis.get('action_plan_alignment', 0) or 0, 0))
            ws.cell(row=row_idx, column=12, value=analisis.get('problema') or '')
            ws.cell(row=row_idx, column=13, value=analisis.get('por_que_1') or '')
            ws.cell(row=row_idx, column=14, value=analisis.get('por_que_2') or '')
            ws.cell(row=row_idx, column=15, value=analisis.get('por_que_3') or '')
            ws.cell(row=row_idx, column=16, value=analisis.get('por_que_4') or '')
            ws.cell(row=row_idx, column=17, value=analisis.get('por_que_5') or '')
            ws.cell(row=row_idx, column=18, value=analisis.get('causa_raiz') or '')
            ws.cell(row=row_idx, column=19, value=analisis.get('plan_accion') or '')
            
            # Feedback IA completo - unir con separador
            ws.cell(row=row_idx, column=20, value=' | '.join(analisis.get('strengths', [])))
            ws.cell(row=row_idx, column=21, value=' | '.join(analisis.get('weaknesses', [])))
            ws.cell(row=row_idx, column=22, value=' | '.join(analisis.get('suggestions', [])))
            ws.cell(row=row_idx, column=23, value=' | '.join(analisis.get('tips', [])))
            ws.cell(row=row_idx, column=24, value=analisis.get('detail') or '')
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 6, 'B': 12, 'C': 10, 'D': 12, 'E': 10, 'F': 12,
            'G': 8, 'H': 8, 'I': 10, 'J': 12, 'K': 12,
            'L': 50, 'M': 50, 'N': 50, 'O': 50, 'P': 50, 'Q': 50,
            'R': 50, 'S': 50, 'T': 60, 'U': 60, 'V': 60, 'W': 60, 'X': 70
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generar nombre de archivo
        filename = 'Analisis_5W'
        if categoria:
            filename += f'_{categoria}'
        if tipo_5w:
            filename += f'_{tipo_5w}'
        if nivel:
            filename += f'_{nivel}'
        if mes:
            filename += f'_{mes}'
        filename += f'_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando a Excel: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# Main
# ============================================================================

if __name__ == '__main__':
    print("🚀 Iniciando Dashboard Admin API...")
    print("📊 Dashboard: http://localhost:5000")
    print("🔌 API Base: http://localhost:5000/api")
    print("\n✅ Endpoints disponibles:")
    print("   GET  /api/analisis")
    print("   POST /api/analisis")
    print("   GET  /api/analisis/:id")
    print("   POST /api/analisis/:id/evaluar")
    print("   GET  /api/estadisticas")
    print("   GET  /api/metricas")
    print("   POST /api/importar")
    print("   GET  /api/exportar-excel")
    print("\n⚡ Presiona Ctrl+C para detener\n")
    
    # 🚂 Railway usa PORT como variable de entorno
    import os
    port = int(os.environ.get('PORT', 5000))
    
    app.run(debug=False, host='0.0.0.0', port=port)
